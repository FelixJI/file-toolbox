"""excel_merge 核心测试:命名、合并语义、样式复制、失败/取消/历史。

全部基于程序化生成的虚构 xlsx(openpyxl 读写),纯 Python、跨平台,不触发 COM。
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from file_toolbox.common.history import JsonHistoryStore
from file_toolbox.core.excel_merge import (
    MODE_FORMULAS,
    MODE_VALUES,
    NAMING_KEEP,
    NAMING_PREFIX,
    ExcelMergeService,
    MergeOptions,
    MergeResult,
    SheetNamer,
    compose_sheet_base,
    sanitize_sheet_name,
)

# ==================== 命名(纯函数) ====================


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("Sheet1", "Sheet1"),  # 合法名原样保留
        ("a[b]c", "a_b_c"),  # 非法字符 [] 替换为 _
        ("a:b\\c/d?e*f", "a_b_c_d_e_f"),  # 全部非法字符逐一替换
        ("   ", "Sheet"),  # 纯空白回退默认名
        ("x" * 40, "x" * 31),  # 超 31 字符截断
        ("  pad  ", "pad"),  # 去首尾空白
    ],
)
def test_sanitize_sheet_name(raw, expected):
    assert sanitize_sheet_name(raw) == expected


def test_compose_sheet_base_by_naming():
    """prefix 策略组合文件主名,keep 策略只用工作表名。"""
    assert compose_sheet_base("报表", "Sheet1", NAMING_PREFIX) == "报表-Sheet1"
    assert compose_sheet_base("报表", "Sheet1", NAMING_KEEP) == "Sheet1"


def test_namer_assign_unique_and_case_insensitive():
    """同名(含大小写不同)冲突时追加 ~n 序号且保持唯一。"""
    namer = SheetNamer()
    first = namer.assign("Data")
    second = namer.assign("DATA")  # Excel 工作表名大小写不敏感,视为冲突
    third = namer.assign("Data")
    assert first == "Data"
    assert second == "DATA~2"
    assert third == "Data~3"
    assert len({first.casefold(), second.casefold(), third.casefold()}) == 3


def test_namer_truncation_keeps_unique_within_31():
    """超长名截断后仍冲突时,缩头加序号,长度不超过 31 且唯一。"""
    namer = SheetNamer()
    long_base = "很长的名字" * 10  # 50 字符
    first = namer.assign(long_base)
    second = namer.assign(long_base)
    assert len(first) == 31
    assert len(second) <= 31
    assert first != second


# ==================== 合并执行 ====================


def test_merge_basic_preserves_order_and_values(make_xlsx, tmp_path):
    """两个文件各两个表 → 输出 4 个工作表,顺序=文件顺序×表顺序,值保留。"""
    a = make_xlsx("a.xlsx", {"Sheet": [["a1"]], "数据": [["x", 1]]})
    b = make_xlsx("b.xlsx", {"Sheet": [["b1"]], "数据": [["y", 2]]})
    out = tmp_path / "merged.xlsx"

    result = ExcelMergeService().merge([a, b], out)

    assert result.success
    assert result.output == out
    assert [m.target_name for m in result.sheets] == [
        "a-Sheet",
        "a-数据",
        "b-Sheet",
        "b-数据",
    ]
    wb = load_workbook(out)
    assert wb.sheetnames == ["a-Sheet", "a-数据", "b-Sheet", "b-数据"]
    assert wb["a-数据"]["A1"].value == "x"
    assert wb["b-数据"]["B1"].value == 2
    assert result.failed == []


def test_merge_creates_missing_output_dir(make_xlsx, tmp_path):
    """输出目录不存在时自动创建(parents=True)。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    out = tmp_path / "not" / "exist" / "merged.xlsx"
    result = ExcelMergeService().merge([a], out)
    assert result.success
    assert result.output.is_file()


def test_merge_keep_naming_conflict_suffix(make_xlsx, tmp_path):
    """keep 命名下跨文件同名工作表自动加序号。"""
    a = make_xlsx("a.xlsx", {"Sheet1": [["a"]]})
    b = make_xlsx("b.xlsx", {"Sheet1": [["b"]]})
    out = tmp_path / "merged.xlsx"

    result = ExcelMergeService().merge([a, b], out, MergeOptions(naming=NAMING_KEEP))

    assert [m.target_name for m in result.sheets] == ["Sheet1", "Sheet1~2"]
    wb = load_workbook(out)
    assert wb.sheetnames == ["Sheet1", "Sheet1~2"]


def test_merge_hidden_sheets(make_xlsx, tmp_path):
    """默认跳过隐藏工作表;include_hidden=True 时并入。"""
    a = make_xlsx("a.xlsx", {"可见": [["v"]]}, extra_sheets={"隐藏表": "hidden"})
    out = tmp_path / "merged.xlsx"
    svc = ExcelMergeService()

    r1 = svc.merge([a], out)
    assert load_workbook(r1.output).sheetnames == ["a-可见"]

    r2 = svc.merge([a], tmp_path / "merged2.xlsx", MergeOptions(include_hidden=True))
    assert load_workbook(r2.output).sheetnames == ["a-可见", "a-隐藏表"]


def test_merge_preserves_styles_and_dimensions(tmp_path):
    """字体加粗、数字格式、合并单元格、列宽、行高均复制到输出。"""
    src = tmp_path / "styled.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "表头"
    ws["A1"].font = Font(bold=True)
    ws["B2"] = 0.25
    ws["B2"].number_format = "0.00%"
    ws.merge_cells("A1:C1")
    ws.column_dimensions["A"].width = 18
    ws.row_dimensions[2].height = 30
    wb.save(src)

    result = ExcelMergeService().merge([src], tmp_path / "out.xlsx")

    m = load_workbook(result.output)["styled-S"]
    assert m["A1"].font.bold is True
    assert m["B2"].number_format == "0.00%"
    assert [str(r) for r in m.merged_cells.ranges] == ["A1:C1"]
    assert m.column_dimensions["A"].width == 18
    assert m.row_dimensions[2].height == 30


def test_merge_formulas_mode_keeps_formula(make_xlsx, tmp_path):
    """formulas 模式保留公式原文。"""
    a = make_xlsx("a.xlsx", {"S": [["=SUM(1,2)"]]})
    r = ExcelMergeService().merge([a], tmp_path / "o.xlsx", MergeOptions(mode=MODE_FORMULAS))
    assert load_workbook(r.output)["a-S"]["A1"].value == "=SUM(1,2)"


def test_merge_values_mode_reads_cache(monkeypatch, make_xlsx, tmp_path):
    """values 模式以 data_only=True 打开(取公式缓存值;无缓存时为空)。"""
    import openpyxl

    a = make_xlsx("a.xlsx", {"S": [["=SUM(1,2)", "plain"]]})
    real = openpyxl.load_workbook
    seen: dict[str, object] = {}

    def spy(path, **kwargs):
        seen.update(kwargs)
        return real(path, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", spy)

    r = ExcelMergeService().merge([a], tmp_path / "o.xlsx", MergeOptions(mode=MODE_VALUES))

    assert seen.get("data_only") is True
    m = load_workbook(r.output)["a-S"]
    # openpyxl 写出的公式无缓存值 → values 模式读到 None(契约:未计算公式为空)
    assert m["A1"].value is None
    assert m["B1"].value == "plain"


def test_merge_failed_source_continues(make_xlsx, tmp_path):
    """损坏文件记入 failed,其余文件照常合并并写出输出。"""
    good = make_xlsx("good.xlsx", {"S": [["v"]]})
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not a zip file")
    r = ExcelMergeService().merge([bad, good], tmp_path / "o.xlsx")

    assert r.success
    assert len(r.failed) == 1
    assert r.failed[0].file == "bad.xlsx"
    assert "无法读取" in r.failed[0].error
    assert [m.target_name for m in r.sheets] == ["good-S"]


def test_merge_all_failed_no_output(tmp_path):
    """全部源文件失败 → 不写输出、success=False、错误消息明确。"""
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"still not a zip")
    r = ExcelMergeService().merge([bad], tmp_path / "o.xlsx")

    assert not r.success
    assert r.output is None
    assert r.error_message == "全部源文件读取失败"
    assert not (tmp_path / "o.xlsx").exists()


def test_merge_unsupported_suffix_clear_error(tmp_path):
    """老格式 .xls 给出明确中文错误(而非 openpyxl 底层报错)。"""
    old = tmp_path / "legacy.xls"
    old.write_bytes(b"PK fake")
    r = ExcelMergeService().merge([old], tmp_path / "o.xlsx")

    assert not r.success
    assert "不支持的格式" in r.failed[0].error


def test_merge_empty_file_list(tmp_path):
    """空文件列表 → 无可合并工作表的失败结果。"""
    r = ExcelMergeService().merge([], tmp_path / "o.xlsx")
    assert not r.success
    assert r.error_message == "没有可合并的工作表"


def test_merge_output_auto_numbering_never_overwrites(make_xlsx, tmp_path):
    """输出已存在时自动加序号,绝不覆盖;非法后缀归一为 .xlsx。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    out = tmp_path / "merged.xlsx"
    out.write_bytes(b"existing precious data")

    r = ExcelMergeService().merge([a], out)

    assert r.output == tmp_path / "merged_1.xlsx"
    assert out.read_bytes() == b"existing precious data"  # 原文件未被覆盖
    assert r.output.is_file()

    # 后缀归一:.xls 输出名被改为 .xlsx
    r2 = ExcelMergeService().merge([a], tmp_path / "again.xls")
    assert r2.output.suffix == ".xlsx"


def test_merge_cancel_between_files(make_xlsx, tmp_path):
    """cancel_check 为真时在下一个文件前取消:不写输出。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    b = make_xlsx("b.xlsx", {"S": [["v"]]})
    out = tmp_path / "o.xlsx"
    calls: list[int] = []

    def cancel_after_first() -> bool:
        return bool(calls)  # 首个文件处理完成后返回 True

    def progress(cur: int, total: int, msg: str) -> None:
        calls.append(cur)

    r = ExcelMergeService().merge(
        [a, b], out, cancel_check=cancel_after_first, progress_callback=progress
    )

    assert r.cancelled
    assert not r.success
    assert not out.exists()
    assert len(r.sheets) == 1  # 首个文件已完成


# ==================== 预览 ====================


def test_plan_sheets(make_xlsx, tmp_path):
    """plan 给出目标名;隐藏表标跳过;损坏文件入 failed。"""
    a = make_xlsx("a.xlsx", {"S1": [["v"]], "S2": [["v"]]}, extra_sheets={"H": "hidden"})
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"nope")

    plans, failed = ExcelMergeService().plan_sheets([a, bad], MergeOptions())

    assert [(p.sheet, p.target_name, p.included) for p in plans] == [
        ("S1", "a-S1", True),
        ("S2", "a-S2", True),
        ("H", "", False),
    ]
    assert plans[2].note == "隐藏工作表"
    assert len(failed) == 1 and failed[0].file == "bad.xlsx"


def test_plan_sheets_hidden_included(make_xlsx):
    """include_hidden=True 时隐藏表也分配目标名并纳入计划。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]}, extra_sheets={"H": "hidden"})
    plans, _ = ExcelMergeService().plan_sheets([a], MergeOptions(include_hidden=True))
    assert [(p.sheet, p.included) for p in plans] == [("S", True), ("H", True)]
    assert plans[1].target_name == "a-H"


# ==================== 历史 ====================


def test_merge_records_history_on_success(make_xlsx, tmp_path):
    """合并成功写入一条 excel_merge 历史,字段与 GUI 历史摘要对应。"""
    history = JsonHistoryStore(history_dir=tmp_path / "h")
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    svc = ExcelMergeService(history_store=history)

    svc.merge([a], tmp_path / "o.xlsx")

    records = history.get_records("excel_merge")
    assert len(records) == 1
    data = records[0]["data"]
    assert data["file_count"] == 1
    assert data["sheet_count"] == 1
    assert data["failed_count"] == 0
    assert data["naming"] == NAMING_PREFIX
    assert data["mode"] == MODE_VALUES
    assert data["success"] is True
    assert data["output"].endswith("o.xlsx")


def test_merge_no_history_on_cancel_or_failure(make_xlsx, tmp_path):
    """取消/全部失败不写历史。"""
    history = JsonHistoryStore(history_dir=tmp_path / "h")
    svc = ExcelMergeService(history_store=history)
    a = make_xlsx("a.xlsx", {"S": [["v"]]})

    r1 = svc.merge([a], tmp_path / "o.xlsx", cancel_check=lambda: True)
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"nope")
    r2 = svc.merge([bad], tmp_path / "o2.xlsx")

    assert r1.cancelled and not r2.success
    assert history.get_records("excel_merge") == []


def test_merge_result_success_semantics():
    """MergeResult.success 只看 output 是否写出;cancelled 恒失败。"""
    assert MergeResult(output=Path("x.xlsx")).success is True
    assert MergeResult(output=None).success is False
    assert MergeResult(output=Path("x.xlsx"), cancelled=True).success is False
