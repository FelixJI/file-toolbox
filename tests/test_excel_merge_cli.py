"""excel-merge CLI 测试:预览默认、--yes 执行、参数校验、失败退出码。

全部基于程序化生成的虚构 xlsx,不触发 COM。
"""

from openpyxl import load_workbook
from typer.testing import CliRunner

from file_toolbox.cli.main import app

runner = CliRunner()


def test_preview_lists_plan_without_writing(make_xlsx, tmp_path):
    """默认预览:列出 工作表 -> 合并后名称,不产生输出文件。"""
    a = make_xlsx("a.xlsx", {"Sheet1": [["v"]], "数据": [["x"]]})
    b = make_xlsx("b.xlsx", {"Sheet1": [["v"]]})

    r = runner.invoke(app, ["excel-merge", str(a), str(b)])

    assert r.exit_code == 0
    assert "a.xlsx | Sheet1 -> a-Sheet1" in r.output
    assert "b.xlsx | Sheet1 -> b-Sheet1" in r.output
    assert "预览模式" in r.output
    assert not (tmp_path / "合并结果.xlsx").exists()


def test_preview_marks_hidden_and_failed(make_xlsx, tmp_path):
    """预览中标出隐藏工作表跳过与读取失败文件。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]}, extra_sheets={"H": "hidden"})
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"nope")

    r = runner.invoke(app, ["excel-merge", str(a), str(bad)])

    assert r.exit_code == 0
    assert "a.xlsx | H [跳过:隐藏工作表]" in r.output
    assert "bad.xlsx [失败]" in r.output


def test_execute_writes_default_output(make_xlsx, tmp_path):
    """--yes 执行:默认输出到首个源文件目录,工作表按文件顺序合并。"""
    a = make_xlsx("a.xlsx", {"S1": [["a"]]})
    b = make_xlsx("b.xlsx", {"S2": [["b"]]})

    r = runner.invoke(app, ["excel-merge", str(a), str(b), "--yes"])

    assert r.exit_code == 0
    out = tmp_path / "合并结果.xlsx"
    assert out.is_file()
    assert load_workbook(out).sheetnames == ["a-S1", "b-S2"]
    assert "完成: 2 个工作表" in r.output


def test_execute_custom_output_and_options(make_xlsx, tmp_path):
    """-o 自定义输出;--naming keep / --mode formulas / --include-hidden 生效。"""
    a = make_xlsx("a.xlsx", {"S": [["=SUM(1,2)"]]}, extra_sheets={"H": "hidden"})
    out = tmp_path / "sub" / "out.pq"  # 非法后缀也应归一为 .xlsx

    r = runner.invoke(
        app,
        [
            "excel-merge",
            str(a),
            "--yes",
            "-o",
            str(out),
            "--naming",
            "keep",
            "--mode",
            "formulas",
            "--include-hidden",
        ],
    )

    assert r.exit_code == 0
    real_out = tmp_path / "sub" / "out.xlsx"
    wb = load_workbook(real_out)
    assert wb.sheetnames == ["S", "H"]
    assert wb["S"]["A1"].value == "=SUM(1,2)"


def test_dir_batch_and_unsupported_warning(make_xlsx, tmp_path):
    """--dir 批量加入:txt 被忽略并提示,仅 xlsx 参与合并。"""
    make_xlsx("a.xlsx", {"S": [["v"]]})
    (tmp_path / "note.txt").write_text("x")

    r = runner.invoke(app, ["excel-merge", "--dir", str(tmp_path), "--yes"])

    assert r.exit_code == 0
    assert "已忽略 1 个不支持的文件" in r.output
    assert (tmp_path / "合并结果.xlsx").is_file()


def test_no_supported_files_errors(tmp_path):
    """只有不支持的文件 → 报错退出 1。"""
    f = tmp_path / "a.txt"
    f.write_text("x")
    r = runner.invoke(app, ["excel-merge", str(f)])
    assert r.exit_code == 1
    assert "未选择任何受支持的 Excel 文件" in r.output


def test_invalid_naming_rejected(make_xlsx):
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    r = runner.invoke(app, ["excel-merge", str(a), "--naming", "bad", "--yes"])
    assert r.exit_code == 1
    assert "无效的 --naming" in r.output


def test_invalid_mode_rejected(make_xlsx):
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    r = runner.invoke(app, ["excel-merge", str(a), "--mode", "bad", "--yes"])
    assert r.exit_code == 1
    assert "无效的 --mode" in r.output


def test_dir_not_a_directory_errors(make_xlsx, tmp_path):
    """--dir 指向文件 → 友好报错而非 traceback。"""
    f = tmp_path / "f.xlsx"
    f.write_bytes(b"")
    r = runner.invoke(app, ["excel-merge", "--dir", str(f)])
    assert r.exit_code == 1
    assert "不是目录" in r.output


def test_execute_all_failed_exits_nonzero(tmp_path):
    """全部源文件失败 → 退出码 1 并给出原因。"""
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"nope")
    r = runner.invoke(app, ["excel-merge", str(bad), "--yes"])
    assert r.exit_code == 1
    assert "全部源文件读取失败" in r.output


def test_output_auto_numbered_when_exists(make_xlsx, tmp_path):
    """输出已存在时自动加序号,原文件不被覆盖。"""
    a = make_xlsx("a.xlsx", {"S": [["v"]]})
    out = tmp_path / "合并结果.xlsx"
    out.write_text("precious")

    r = runner.invoke(app, ["excel-merge", str(a), "--yes", "-o", str(out)])

    assert r.exit_code == 0
    assert out.read_text(encoding="utf-8") == "precious"
    assert (tmp_path / "合并结果_1.xlsx").is_file()
