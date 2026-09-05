"""ExcelMergeWorker 测试:正常完成、异常、进度、取消、真实合并集成。

worker.run() 同步调用(不走 QThread.start)验证逻辑;另含真实 start() 的
跨线程信号投递集成(与 test_invoice_worker 同范式)。合并基于虚构 xlsx,不触发 COM。
"""

import pytest

# 用 QtWidgets 子模块做 importorskip(而非顶层 PySide6):后者只校验包可 import,
# 不触发 libEGL/libGL 原生库加载;真实 import QtWidgets 才会,缺库时应跳过而非收集失败。
pytest.importorskip("PySide6.QtWidgets")

from pathlib import Path  # noqa: E402

from openpyxl import load_workbook  # noqa: E402
from PySide6.QtWidgets import QApplication  # noqa: E402

from file_toolbox.core.excel_merge import (  # noqa: E402
    ExcelMergeService,
    MergeOptions,
    MergeResult,
)
from file_toolbox.gui.workers.excel_merge_worker import ExcelMergeWorker  # noqa: E402


@pytest.fixture(scope="module")
def app():
    return QApplication.instance() or QApplication([])


class _FakeService:
    """假 ExcelMergeService:记录调用,可控返回/异常。"""

    def __init__(self, result=None, error=None):
        self._result = result
        self._error = error
        self.merge_calls: list[dict] = []

    def merge(self, files, output, options, progress_callback=None, cancel_check=None):
        self.merge_calls.append({"files": list(files), "output": output, "options": options})
        if self._error:
            raise self._error
        if progress_callback is not None:
            for i in range(len(files)):
                progress_callback(i + 1, len(files), f"合并 {files[i].name}")
        return self._result


def test_worker_emits_finished_ok_on_success(app):
    """正常完成 → finished_ok 信号带 MergeResult,透传 files/output/options。"""
    expected = MergeResult(output=Path("o.xlsx"))
    svc = _FakeService(result=expected)
    worker = ExcelMergeWorker(svc, [Path("a.xlsx"), Path("b.xlsx")], Path("o.xlsx"), MergeOptions())

    captured = {}
    worker.finished_ok.connect(lambda r: captured.setdefault("ok", r))
    worker.failed.connect(lambda m: captured.setdefault("fail", m))

    worker.run()  # 同步跑(不经 QThread.start)

    assert captured.get("ok") is expected
    assert "fail" not in captured
    assert len(svc.merge_calls) == 1
    assert svc.merge_calls[0]["output"] == Path("o.xlsx")


def test_worker_emits_failed_on_exception(app):
    """service 抛异常 → failed 信号,不发 finished_ok。"""
    svc = _FakeService(error=RuntimeError("boom"))
    worker = ExcelMergeWorker(svc, [Path("a.xlsx")], Path("o.xlsx"), MergeOptions())

    captured = {}
    worker.finished_ok.connect(lambda r: captured.setdefault("ok", r))
    worker.failed.connect(lambda m: captured.setdefault("fail", m))

    worker.run()

    assert "boom" in captured.get("fail", "")
    assert "ok" not in captured


def test_worker_emits_progress(app):
    """进度回调 → progress 信号,(current, total, msg) 形状正确。"""
    svc = _FakeService(result=MergeResult())
    worker = ExcelMergeWorker(
        svc, [Path(f"f{i}.xlsx") for i in range(3)], Path("o.xlsx"), MergeOptions()
    )

    progress_msgs = []
    worker.progress.connect(lambda c, t, m: progress_msgs.append((c, t, m)))

    worker.run()

    assert progress_msgs == [
        (1, 3, "合并 f0.xlsx"),
        (2, 3, "合并 f1.xlsx"),
        (3, 3, "合并 f2.xlsx"),
    ]


def test_worker_cancel_sets_flag(app):
    """cancel() 设标志;_cancel_check 反映该标志。"""
    svc = _FakeService(result=MergeResult())
    worker = ExcelMergeWorker(svc, [Path("a.xlsx")], Path("o.xlsx"), MergeOptions())

    assert worker._cancel is False
    assert worker._cancel_check() is False
    worker.cancel()
    assert worker._cancel is True
    assert worker._cancel_check() is True


def test_worker_start_merges_real_files_across_threads(app, make_xlsx, tmp_path):
    """集成:真实 service + worker.start() 后台线程 → 合并结果投递回主线程。"""
    a = make_xlsx("a.xlsx", {"S1": [["v"]]})
    b = make_xlsx("b.xlsx", {"S2": [["v"]]})
    out = tmp_path / "合并结果.xlsx"
    svc = ExcelMergeService()
    worker = ExcelMergeWorker(svc, [a, b], out, MergeOptions())

    results: list[MergeResult] = []
    worker.finished_ok.connect(results.append)
    failed: list[str] = []
    worker.failed.connect(failed.append)

    worker.start()
    assert worker.wait(10000), "worker 未在 10s 内结束"
    app.processEvents()
    if not results:
        app.processEvents()

    assert failed == []
    assert len(results) == 1 and results[0].success
    assert load_workbook(out).sheetnames == ["a-S1", "b-S2"]
