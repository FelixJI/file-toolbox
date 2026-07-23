import json

from openpyxl import load_workbook

from file_toolbox.core.invoice.exporters.excel_exporter import export_excel
from file_toolbox.core.invoice.types import Invoice, LineItem


def _invoice(num="A", dup=False) -> Invoice:
    return Invoice(
        invoice_number=num,
        invoice_type="增值税专用发票",
        issue_date="2026-05-19",
        seller_name="测试销售方",
        seller_tax_id="STAX",
        seller_addr="sa",
        seller_tel="st",
        seller_bank="sb",
        seller_account="sac",
        buyer_name="测试购买方",
        buyer_tax_id="BTAX",
        buyer_addr="ba",
        buyer_tel="bt",
        buyer_bank="bb",
        buyer_account="bac",
        amount_without_tax="100.00",
        tax_amount="13.00",
        amount_with_tax="113.00",
        amount_chinese="壹佰壹拾叁圆整",
        drawer="测试人",
        remark="备注",
        items=[LineItem("*交通运输设备*甲", "S1", "根", "1", "100", "100.00", "13%", "13.00")],
        source_file="x.xml",
        parse_method="xml",
        is_duplicate=dup,
    )


def test_excel_two_sheets(tmp_path):
    out = tmp_path / "out.xlsx"
    export_excel([_invoice("A"), _invoice("B")], out)
    wb = load_workbook(out)
    assert "发票汇总" in wb.sheetnames
    assert "明细清单" in wb.sheetnames


def test_excel_summary_headers_and_row(tmp_path):
    out = tmp_path / "out.xlsx"
    export_excel([_invoice("A")], out)
    wb = load_workbook(out)
    ws = wb["发票汇总"]
    headers = [c.value for c in ws[1]]
    assert "发票号码" in headers
    assert "解析方式" in headers
    # 数据行
    row2 = [c.value for c in ws[2]]
    assert row2[0] == "A"
    assert "测试销售方" in row2


def test_excel_detail_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    export_excel([_invoice("A")], out)
    wb = load_workbook(out)
    ws = wb["明细清单"]
    headers = [c.value for c in ws[1]]
    assert "发票号码" in headers
    assert "项目名称" in headers
    row2 = [c.value for c in ws[2]]
    assert row2[0] == "A"  # 外键
    assert row2[1] == "*交通运输设备*甲"


def test_excel_duplicate_marked_yellow(tmp_path):
    out = tmp_path / "out.xlsx"
    export_excel([_invoice("A", dup=True), _invoice("B", dup=False)], out)
    wb = load_workbook(out)
    ws = wb["发票汇总"]
    # 重复行(第2行)填充色应为黄色族
    fill = ws.cell(row=2, column=1).fill
    assert fill.patternType == "solid"
    assert str(fill.fgColor.rgb).endswith("FFF2CC")


def test_json_export_structure(tmp_path):
    from file_toolbox.core.invoice.exporters.json_exporter import export_json
    from file_toolbox.core.invoice.types import FailedFile

    out = tmp_path / "out.json"
    invs = [_invoice("A")]
    failed = [FailedFile(file="bad.zip", reason="损坏")]
    export_json(invs, out, dedupe_strategy="mark", failed=failed)

    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["dedupe_strategy"] == "mark"
    assert "exported_at" in data
    assert len(data["invoices"]) == 1
    inv0 = data["invoices"][0]
    assert inv0["invoice_number"] == "A"
    assert inv0["seller"]["name"] == "测试销售方"
    assert inv0["buyer"]["name"] == "测试购买方"
    assert inv0["items"][0]["name"] == "*交通运输设备*甲"
    assert inv0["parse_method"] == "xml"
    assert data["failed"][0]["file"] == "bad.zip"


def test_json_chinese_not_escaped(tmp_path):
    """JSON 中文应为明文,不转义为 \\uXXXX。"""
    from file_toolbox.core.invoice.exporters.json_exporter import export_json

    out = tmp_path / "out.json"
    export_json([_invoice("A")], out, dedupe_strategy="keep_all")
    raw = out.read_text(encoding="utf-8")
    assert "测试销售方" in raw  # 中文明文,未转义
