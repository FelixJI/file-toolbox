"""Excel 双 Sheet 导出。Sheet1 发票汇总(一行一发票),Sheet2 明细清单(一行一明细)。

重复行(is_duplicate=True)标浅黄底。openpyxl 延迟导入,缺依赖给友好提示。
"""

from pathlib import Path

from file_toolbox.core.invoice.types import Invoice

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    raise ImportError("Excel 导出需要 openpyxl: pip install 'file-toolbox[invoice]'") from e


_DUP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Sheet1 列定义
_SUMMARY_HEADERS = [
    "发票号码",
    "发票类型",
    "开票日期",
    "销售方名称",
    "销售方税号",
    "购买方名称",
    "购买方税号",
    "不含税金额",
    "税额",
    "价税合计",
    "大写金额",
    "开票人",
    "备注",
    "来源文件",
    "解析方式",
]

# Sheet2 列定义
_DETAIL_HEADERS = [
    "发票号码",
    "项目名称",
    "规格型号",
    "单位",
    "数量",
    "单价",
    "金额",
    "税率",
    "税额",
]


def _summary_row(inv: Invoice) -> list[str]:
    return [
        inv.invoice_number,
        inv.invoice_type,
        inv.issue_date,
        inv.seller_name,
        inv.seller_tax_id,
        inv.buyer_name,
        inv.buyer_tax_id,
        inv.amount_without_tax,
        inv.tax_amount,
        inv.amount_with_tax,
        inv.amount_chinese,
        inv.drawer,
        inv.remark,
        inv.source_file,
        inv.parse_method,
    ]


def _apply_fill(ws: Worksheet, row_idx: int, ncols: int, fill: PatternFill) -> None:
    for col in range(1, ncols + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def export_excel(invoices: list[Invoice], output_path: Path) -> Path:
    """导出为双 Sheet Excel。返回输出路径。"""
    output_path = Path(output_path)
    wb = Workbook()

    # Sheet1 发票汇总
    ws1 = wb.active
    ws1.title = "发票汇总"
    ws1.append(_SUMMARY_HEADERS)
    for inv in invoices:
        ws1.append(_summary_row(inv))
        if inv.is_duplicate:
            _apply_fill(ws1, ws1.max_row, len(_SUMMARY_HEADERS), _DUP_FILL)

    # Sheet2 明细清单
    ws2 = wb.create_sheet("明细清单")
    ws2.append(_DETAIL_HEADERS)
    for inv in invoices:
        for item in inv.items:
            ws2.append(
                [
                    inv.invoice_number,
                    item.name,
                    item.spec,
                    item.unit,
                    item.quantity,
                    item.unit_price,
                    item.amount,
                    item.tax_rate,
                    item.tax_amount,
                ]
            )
            if inv.is_duplicate:
                _apply_fill(ws2, ws2.max_row, len(_DETAIL_HEADERS), _DUP_FILL)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
