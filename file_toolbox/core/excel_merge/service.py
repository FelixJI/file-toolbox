"""Excel 合并核心:多个工作簿的工作表复制进一个新工作簿。

纯 openpyxl 实现,跨平台、不依赖 Office/WPS COM。openpyxl 在方法内延迟导入,
避免 CLI 入口与 GUI 启动链为 `--help`/首屏付出 openpyxl 导入成本(与 invoice
导出同策略)。复制内容:单元格值(或公式)、数字格式、字体/填充/边框/保护/对齐、
合并单元格、列宽与行高;图表、图片、透视表与宏(xlsm 的 VBA)不复制。
"""

from __future__ import annotations

from collections.abc import Callable
from copy import copy
from pathlib import Path
from typing import TYPE_CHECKING

from file_toolbox.common.history import JsonHistoryStore
from file_toolbox.common.loggable import LoggableMixin
from file_toolbox.core.excel_merge.constants import (
    MODE_VALUES,
    SUPPORTED_SUFFIXES,
)
from file_toolbox.core.excel_merge.naming import SheetNamer, compose_sheet_base
from file_toolbox.core.excel_merge.types import (
    FailedSource,
    MergedSheet,
    MergeOptions,
    MergeResult,
    SheetPlan,
)

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

ProgressCallback = Callable[[int, int, str], None]
CancelCheck = Callable[[], bool]


class ExcelMergeService(LoggableMixin):
    """Excel 合并服务:plan_sheets 预览、merge 执行。"""

    def __init__(self, history_store: JsonHistoryStore | None = None) -> None:
        """初始化服务。

        Args:
            history_store: 历史存储;传入则 merge 成功写出输出后记录一条
                excel_merge 历史。None 表示不记录(默认)。
        """
        self._history_store = history_store

    # ==================== 预览 ====================

    def plan_sheets(
        self,
        files: list[Path],
        options: MergeOptions | None = None,
        progress_callback: ProgressCallback | None = None,
    ) -> tuple[list[SheetPlan], list[FailedSource]]:
        """读取每个源文件的工作表清单,给出合并后的目标名(不复制内容)。

        Args:
            files: 源文件列表。
            options: 合并选项(命名策略决定目标名);None 用默认(prefix/values)。
            progress_callback: (current, total, message) 进度回调。

        Returns:
            (计划行列表, 读取失败列表)。included=False 的行不会进入输出。
        """
        if options is None:
            options = MergeOptions()
        plans: list[SheetPlan] = []
        failed: list[FailedSource] = []
        namer = SheetNamer()
        total = len(files)
        for idx, path in enumerate(files, start=1):
            if progress_callback is not None:
                progress_callback(idx, total, f"读取 {path.name}")
            try:
                src = self._load_workbook(path, MODE_VALUES)
            except Exception as e:
                failed.append(FailedSource(path.name, f"无法读取: {e}"))
                continue
            for ws in src.worksheets:
                if ws.sheet_state != "visible" and not options.include_hidden:
                    plans.append(SheetPlan(path.name, ws.title, "", False, "隐藏工作表"))
                    continue
                target = namer.assign(compose_sheet_base(path.stem, ws.title, options.naming))
                plans.append(SheetPlan(path.name, ws.title, target, True))
            src.close()
        return plans, failed

    # ==================== 执行 ====================

    def merge(
        self,
        files: list[Path],
        output: Path,
        options: MergeOptions | None = None,
        progress_callback: ProgressCallback | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> MergeResult:
        """把 files 的工作表依次合并进一个新建工作簿并写出。

        Args:
            files: 源文件列表(按顺序合并)。
            output: 输出文件(后缀强制 .xlsx;已存在时自动加序号,绝不覆盖)。
            options: 合并选项;None 用默认(prefix/values/不含隐藏表)。
            progress_callback: (current, total, message) 进度回调(按源文件粒度)。
            cancel_check: 返回 True 时在下一个源文件前取消(不写输出)。

        Returns:
            MergeResult:success = 输出已写出(允许部分源文件失败)。
        """
        if options is None:
            options = MergeOptions()
        dest = self._new_workbook()
        namer = SheetNamer()
        merged: list[MergedSheet] = []
        failed: list[FailedSource] = []
        cancelled = False
        total = len(files)

        for idx, path in enumerate(files, start=1):
            if cancel_check is not None and cancel_check():
                cancelled = True
                self.logger.info("Excel 合并被取消,已处理 %d/%d 个文件", idx - 1, total)
                break
            if progress_callback is not None:
                progress_callback(idx, total, f"合并 {path.name}")
            try:
                src = self._load_workbook(path, options.mode)
            except Exception as e:
                failed.append(FailedSource(path.name, f"无法读取: {e}"))
                self.logger.warning("源文件读取失败,跳过: %s (%s)", path, e)
                continue
            for ws in src.worksheets:
                if ws.sheet_state != "visible" and not options.include_hidden:
                    continue
                target = namer.assign(compose_sheet_base(path.stem, ws.title, options.naming))
                self._copy_worksheet(ws, dest, target)
                merged.append(MergedSheet(path.name, ws.title, target))
            src.close()

        if cancelled:
            return MergeResult(output=None, sheets=merged, failed=failed, cancelled=True)
        if not merged:
            reason = (
                "全部源文件读取失败" if failed and len(failed) == total else "没有可合并的工作表"
            )
            return MergeResult(output=None, sheets=[], failed=failed, error_message=reason)

        output_path = self._resolve_output_path(self._normalize_output(output))
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            dest.save(output_path)
        except Exception as e:
            self.logger.error("输出工作簿写出失败: %s (%s)", output_path, e)
            return MergeResult(
                output=None, sheets=merged, failed=failed, error_message=f"输出失败: {e}"
            )
        self.logger.info(
            "Excel 合并完成: %d 个文件 -> %d 个工作表 -> %s", total, len(merged), output_path
        )
        result = MergeResult(output=output_path, sheets=merged, failed=failed)
        self._record_history(result, total, options)
        return result

    # ==================== 内部实现 ====================

    def _load_workbook(self, path: Path, mode: str) -> Workbook:
        """打开源工作簿。后缀先于 openpyxl 校验,给清晰的中文错误。"""
        if path.suffix.lower() not in SUPPORTED_SUFFIXES:
            raise ValueError(
                f"不支持的格式 {path.suffix or '(无后缀)'},仅支持 {'/'.join(SUPPORTED_SUFFIXES)}"
            )
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover -- openpyxl 是 base 依赖,测试环境必装
            raise ImportError("Excel 合并需要 openpyxl 依赖,请重新安装 file-toolbox") from e
        return load_workbook(path, data_only=mode == MODE_VALUES)

    def _new_workbook(self) -> Workbook:
        """新建空输出工作簿(移除默认空工作表,避免占用 "Sheet" 名称)。"""
        from openpyxl import Workbook

        dest = Workbook()
        dest.remove(dest.active)
        return dest

    def _copy_worksheet(self, src_ws: Worksheet, dest: Workbook, title: str) -> Worksheet:
        """把源工作表的内容与样式复制为输出工作簿中的新工作表。"""
        dst = dest.create_sheet(title=title)
        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is None and not cell.has_style:
                    continue
                dst_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    # openpyxl 的样式对象属于源工作簿,需 copy 出独立副本再赋给目标
                    dst_cell.font = copy(cell.font)
                    dst_cell.border = copy(cell.border)
                    dst_cell.fill = copy(cell.fill)
                    dst_cell.number_format = cell.number_format
                    dst_cell.protection = copy(cell.protection)
                    dst_cell.alignment = copy(cell.alignment)
        for rng in src_ws.merged_cells.ranges:
            dst.merge_cells(str(rng))
        for letter, dim in src_ws.column_dimensions.items():
            if dim.width is not None:
                dst.column_dimensions[letter].width = dim.width
        for index, dim in src_ws.row_dimensions.items():
            if dim.height is not None:
                dst.row_dimensions[index].height = dim.height
        return dst

    def _normalize_output(self, output: Path) -> Path:
        """输出统一为 .xlsx(合并不保留 xlsm 宏)。"""
        if output.suffix.lower() != ".xlsx":
            return output.with_suffix(".xlsx")
        return output

    def _resolve_output_path(self, output: Path) -> Path:
        """输出已存在时自动加序号,绝不覆盖已有文件(与 pdf 输出策略一致)。"""
        if not output.exists():
            return output
        counter = 1
        while True:
            candidate = output.with_name(f"{output.stem}_{counter}{output.suffix}")
            if not candidate.exists():
                return candidate
            counter += 1

    def _record_history(self, result: MergeResult, file_count: int, options: MergeOptions) -> None:
        """记录 excel_merge 历史(若注入了 history_store 且合并成功)。"""
        if self._history_store is None or not result.success:
            return
        assert result.output is not None  # success 蕴含 output 已写出
        self._history_store.add_record(
            "excel_merge",
            {
                "output": str(result.output),
                "file_count": file_count,
                "sheet_count": len(result.sheets),
                "failed_count": len(result.failed),
                "naming": options.naming,
                "mode": options.mode,
                "success": True,
            },
        )
